#!/usr/bin/env python
# encoding: utf-8
'''
@author: toby
@license: (C) Copyright 2013-2017, Node Supply Chain Manager Corporation Limited.
@contact: hejunwang01@gmail.com
@software: garner
@file: opera_exl.py
@time: 2020/11/13 13:44
@desc:
'''

'''

'''

from openpyxl import load_workbook

import os

class Read_exl:

    def __init__(self,filename,param = None):
        '''
        初始化
        :param filename:
        :param param:
        '''
        self.wb = load_workbook(filename)
        if param is None :
            self.tb = self.wb['Sheet1']
        if param:
            self.tb =self.wb[param]

        print('所有sheetname:{}'.format(self.wb.get_sheet_names()))   #所有的sheetname
        print('当前sheetname:{}'.format(self.tb))   #激活 sheet1表格
        print('max_row-->{}'.format(self.tb.max_row))   #sheet表格中的总行数
        print('max_column-->{}'.format(self.tb.max_column))  # l列数

    @property
    def readexcel(self):
        '''
        读取文件内容,存到list中
        :return:
        '''
        list = []
        for i in range(1,self.tb.max_row):
            list2 = []
            for j in range(1,self.tb.max_column):
                # print(self.tb.cell(row = i,column = j).value)
                list2.append(self.tb.cell(row = i,column = j).value)

            list.append(list2)

        for i in list:
            print(i)


        return list



    def writeexl(self,row,column,text):
        '''
        写入文件内容
        :param row:
        :param column:
        :param text:
        :return:
        '''
        self.tb.cell(row=row,column=column).value=text

        self.wb.save('../data/dd.xlsx')





if __name__ == '__main__':
    path_ = os.path.abspath(os.path.dirname(os.getcwd()) + os.path.sep + ".") + '\data'
    # Read_exl(os.path.join(path_,'readcsv.xlsx'))
    print(path_)
    # ls=Read_exl('../data/dd.xlsx').readexcel()

    Read_exl('../data/dd.xlsx','Sheet3').writeexl(1,5,'result')

    ll = Read_exl('../data/dd.xlsx','Sheet2').readexcel
    print('\n')
    for i in ll:
        print(i)